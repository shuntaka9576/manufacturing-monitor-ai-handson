"""製造設備モニタリング用サンプルExcelデータ生成スクリプト"""

import math
import random
import shutil
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

random.seed(42)

BASE_DIR = Path(__file__).parent.parent.parent
OUTPUT_DIRS = [
    BASE_DIR / "ch1",
    BASE_DIR / "ch2",
    BASE_DIR / "ch3",
    BASE_DIR / "ch4",
    BASE_DIR / "ch4_fin",
]

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

EQUIPMENT_NAMES = {
    1: "CNC旋盤 A-01",
    2: "CNC旋盤 A-02",
    3: "プレス機 B-01",
    4: "プレス機 B-02",
    5: "射出成形機 C-01",
    6: "射出成形機 C-02",
    7: "溶接ロボット D-01",
    8: "溶接ロボット D-02",
}

EQUIPMENT_TYPES = {
    1: "CNC旋盤",
    2: "CNC旋盤",
    3: "プレス機",
    4: "プレス機",
    5: "射出成形機",
    6: "射出成形機",
    7: "溶接ロボット",
    8: "溶接ロボット",
}

# センサープロファイル: {設備タイプ: {パラメータ: (mean, std)}}
SENSOR_PROFILES = {
    "CNC旋盤": {
        "temperature": (40, 8),
        "vibration": (2.5, 0.8),
        "rpm": (1800, 500),
        "power_kw": (10, 3),
    },
    "プレス機": {
        "temperature": (38, 6),
        "vibration": (5.0, 1.5),
        "power_kw": (35, 8),
        "pressure": (20, 5),
    },
    "射出成形機": {
        "temperature": (200, 10),
        "vibration": (1.8, 0.6),
        "power_kw": (25, 5),
        "pressure": (100, 25),
    },
    "溶接ロボット": {
        "temperature": (55, 10),
        "vibration": (3.0, 1.0),
        "power_kw": (18, 4),
    },
}

SENSOR_PARAMS = ["temperature", "vibration", "rpm", "power_kw", "pressure"]

# 異常パターン定義: {設備ID: {pattern_type, hours, params, rate}}
ANOMALY_PATTERNS = {
    6: {"pattern_type": "anomaly_rise", "hours": 3, "params": ["temperature", "vibration"], "rate": 0.15},
    4: {"pattern_type": "zero_fill", "hours": 6, "params": "all", "rate": None},
    8: {"pattern_type": "low_value", "hours": 3, "params": "all", "rate": None},
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def create_equipment_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "設備マスタ"
    ws.append(["設備名", "タイプ", "設置場所", "設置日"])
    data = [
        ("CNC旋盤 A-01", "CNC旋盤", "A棟1F", "2020-04-15"),
        ("CNC旋盤 A-02", "CNC旋盤", "A棟1F", "2020-06-20"),
        ("プレス機 B-01", "プレス機", "B棟1F", "2019-11-10"),
        ("プレス機 B-02", "プレス機", "B棟1F", "2021-01-25"),
        ("射出成形機 C-01", "射出成形機", "A棟2F", "2021-08-05"),
        ("射出成形機 C-02", "射出成形機", "A棟2F", "2022-03-12"),
        ("溶接ロボット D-01", "溶接ロボット", "B棟2F", "2022-07-01"),
        ("溶接ロボット D-02", "溶接ロボット", "B棟2F", "2023-02-14"),
    ]
    for row in data:
        ws.append(row)
    style_header(ws)
    auto_width(ws)


def create_status_log_sheet(wb: Workbook, now: datetime):
    ws = wb.create_sheet("ステータス変更履歴")
    ws.append(["設備ID", "設備名", "発生日時", "変更前ステータス", "変更後ステータス", "理由"])

    # fmt: off
    events = [
        # === 異常・故障系 ===
        (6, 7.0, "稼働中", "異常", "冷却水温度が上限値を超過"),
        (6, 6.8, "異常", "メンテナンス中", "冷却水系統の点検開始"),
        (6, 6.5, "メンテナンス中", "稼働中", "冷却水ポンプ交換完了、稼働再開"),
        (7, 6.5, "稼働中", "異常", "ベアリング摩耗による振動増加を検知"),
        (7, 6.0, "異常", "メンテナンス中", "ベアリング交換作業開始"),
        (7, 5.5, "メンテナンス中", "稼働中", "ベアリング交換完了、稼働再開"),
        (3, 6.0, "稼働中", "異常", "油圧圧力が規定値を下回る低下を検知"),
        (3, 5.8, "異常", "メンテナンス中", "油圧系統の修理開始"),
        (3, 5.5, "メンテナンス中", "稼働中", "油圧ポンプ修理完了、稼働再開"),
        (1, 5.5, "稼働中", "異常", "サーボモーター過負荷により自動停止"),
        (1, 5.0, "異常", "メンテナンス中", "サーボモーター点検開始"),
        (1, 4.5, "メンテナンス中", "稼働中", "モータードライバ交換完了、稼働再開"),
        (2, 5.0, "稼働中", "停止中", "電源電圧変動により安全停止"),
        (2, 4.8, "停止中", "稼働中", "電源安定化後、稼働再開"),
        (5, 4.5, "稼働中", "異常", "冷却ファン故障によるヒーター温度上昇"),
        (5, 4.0, "異常", "メンテナンス中", "冷却ファン交換作業開始"),
        (5, 3.8, "メンテナンス中", "稼働中", "冷却ファン交換完了、稼働再開"),
        (8, 4.0, "稼働中", "異常", "エンコーダ信号異常を検知"),
        (8, 3.5, "異常", "メンテナンス中", "エンコーダ交換作業開始"),
        (8, 3.0, "メンテナンス中", "停止中", "エンコーダ交換完了、動作確認待ち"),
        # === 設備部品系 ===
        (1, 4.0, "稼働中", "停止中", "潤滑油劣化による予防停止"),
        (1, 3.8, "停止中", "メンテナンス中", "潤滑油交換作業開始"),
        (1, 3.5, "メンテナンス中", "稼働中", "潤滑油交換完了、稼働再開"),
        (2, 1.5, "稼働中", "停止中", "切削工具摩耗による交換停止"),
        (2, 1.3, "停止中", "稼働中", "工具交換完了、稼働再開"),
        (7, 1.0, "稼働中", "異常", "ロボットアーム位置ずれを検知"),
        (7, 0.8, "異常", "メンテナンス中", "ロボットアーム校正作業開始"),
        (7, 0.6, "メンテナンス中", "稼働中", "ロボットアーム校正完了、稼働再開"),
        (1, 0.8, "稼働中", "異常", "スピンドルモーター異常振動を検知"),
        (1, 0.6, "異常", "停止中", "スピンドルモーター点検のため停止"),
        (1, 0.4, "停止中", "稼働中", "スピンドルベアリング交換完了、稼働再開"),
        # === 品質・生産系 ===
        (5, 3.5, "稼働中", "停止中", "原材料品質不良によるロット停止"),
        (5, 3.3, "停止中", "稼働中", "原材料交換後、稼働再開"),
        (2, 3.0, "稼働中", "停止中", "加工精度低下により調整停止"),
        (2, 2.8, "停止中", "稼働中", "工具交換・調整完了、稼働再開"),
        (3, 3.0, "稼働中", "停止中", "ロット切替のため段取り替え停止"),
        (3, 2.5, "停止中", "稼働中", "段取り替え完了、稼働再開"),
        (6, 2.5, "稼働中", "停止中", "試作品用の設定変更のため停止"),
        (6, 2.3, "停止中", "稼働中", "試作品設定完了、稼働再開"),
        # === 計画・メンテナンス系 ===
        (4, 5.0, "稼働中", "メンテナンス中", "年次法定点検のため計画停止"),
        (4, 4.0, "メンテナンス中", "稼働中", "年次法定点検完了、稼働再開"),
        (3, 2.0, "稼働中", "メンテナンス中", "新規金型セットアップのため停止"),
        (3, 1.8, "メンテナンス中", "稼働中", "金型セットアップ完了、稼働再開"),
        (7, 2.0, "稼働中", "メンテナンス中", "フィルター交換のため計画停止"),
        (7, 1.8, "メンテナンス中", "稼働中", "フィルター交換完了、稼働再開"),
        (1, 1.5, "稼働中", "メンテナンス中", "制御ソフトウェアアップデートのため停止"),
        (1, 1.3, "メンテナンス中", "稼働中", "ソフトウェアアップデート完了、稼働再開"),
        (8, 1.5, "停止中", "メンテナンス中", "安全装置動作確認のため点検"),
        (8, 1.2, "メンテナンス中", "停止中", "安全装置点検完了、次回起動待ち"),
        # === 外的要因系 ===
        (4, 1.0, "稼働中", "停止中", "停電発生により全設備緊急停止"),
        (4, 0.9, "停止中", "メンテナンス中", "停電復旧後の設備点検開始"),
        (4, 0.8, "メンテナンス中", "稼働中", "停電復旧点検完了、稼働再開"),
        # === 通常運用系 ===
        (3, 1.0, "稼働中", "停止中", "材料補充のため一時停止"),
        (3, 0.8, "停止中", "稼働中", "材料補充完了、稼働再開"),
        (5, 0.5, "稼働中", "停止中", "原料切れによる停止"),
        (5, 0.3, "停止中", "稼働中", "原料補充完了、稼働再開"),
        # === 最終ステータスに至るイベント ===
        # 設備1 (CNC旋盤 A-01): 稼働中 — 上記で稼働再開済み
        # 設備2 (CNC旋盤 A-02): 稼働中 — 上記で稼働再開済み
        # 設備3 (プレス機 B-01): 稼働中 — 上記で稼働再開済み
        # 設備4 (プレス機 B-02): メンテナンス中
        (4, 0.5, "稼働中", "メンテナンス中", "定期メンテナンス開始"),
        # 設備5 (射出成形機 C-01): 稼働中 — 上記で稼働再開済み
        # 設備6 (射出成形機 C-02): 異常
        (6, 1.0, "稼働中", "異常", "温度異常上昇を検知"),
        # 設備7 (溶接ロボット D-01): 稼働中 — 上記で稼働再開済み
        # 設備8 (溶接ロボット D-02): 停止中
        (8, 0.3, "停止中", "停止中", "受注減による生産調整で停止継続"),
    ]
    # fmt: on

    for equip_id, days_ago, old_s, new_s, reason in events:
        ts = now - timedelta(days=days_ago)
        ws.append((equip_id, EQUIPMENT_NAMES[equip_id], ts.isoformat(timespec="seconds"), old_s, new_s, reason))
    style_header(ws)
    auto_width(ws)


def create_sensor_data_sheet(wb: Workbook, now: datetime):
    """センサー時系列データを生成してシートに格納する"""
    ws = wb.create_sheet("センサーデータ")
    ws.append(["設備ID", "タイムスタンプ", "temperature", "vibration", "rpm", "power_kw", "pressure"])

    for equip_id in range(1, 9):
        equip_type = EQUIPMENT_TYPES[equip_id]
        profile = SENSOR_PROFILES.get(equip_type, {})
        anomaly = ANOMALY_PATTERNS.get(equip_id)

        for i in range(144):  # 24h × 6 (10分間隔)
            ts = now - timedelta(minutes=(143 - i) * 10)
            hours_ago = (now - ts).total_seconds() / 3600

            # 異常パターン: zero_fill（メンテナンス中の設備は全センサー値ゼロ）
            if anomaly and anomaly["pattern_type"] == "zero_fill" and hours_ago < anomaly["hours"]:
                ws.append((equip_id, ts.isoformat(timespec="seconds"), 0, 0, 0, 0, 0))
                continue

            # 異常パターン: low_value（停止中の設備は低値）
            if anomaly and anomaly["pattern_type"] == "low_value" and hours_ago < anomaly["hours"]:
                ws.append((equip_id, ts.isoformat(timespec="seconds"), 20, 0.1, 0, 0.5, 0))
                continue

            values = {}
            for param in SENSOR_PARAMS:
                spec = profile.get(param)
                if spec is None:
                    values[param] = None
                    continue

                base, std = spec
                daily_cycle = math.sin(2 * math.pi * (ts.hour / 24)) * std * 0.3
                value = random.gauss(base, std * 0.5) + daily_cycle

                # 異常パターン: anomaly_rise（段階的にセンサー値が上昇）
                if anomaly and anomaly["pattern_type"] == "anomaly_rise" and hours_ago < anomaly["hours"]:
                    target_params = anomaly["params"]
                    if target_params == "all" or param in target_params:
                        anomaly_factor = 1 + (anomaly["hours"] - hours_ago) * anomaly["rate"]
                        value *= anomaly_factor

                values[param] = round(max(0, value), 2)

            ws.append((
                equip_id,
                ts.isoformat(timespec="seconds"),
                values.get("temperature"),
                values.get("vibration"),
                values.get("rpm"),
                values.get("power_kw"),
                values.get("pressure"),
            ))

    style_header(ws)
    auto_width(ws)


def main():
    now = datetime.now()

    wb = Workbook()
    create_equipment_sheet(wb)
    create_status_log_sheet(wb, now)
    create_sensor_data_sheet(wb, now)

    primary = OUTPUT_DIRS[0] / "sample_data.xlsx"
    wb.save(str(primary))
    print(f"Excelファイルを生成しました: {primary}")

    for output_dir in OUTPUT_DIRS[1:]:
        output_path = output_dir / "sample_data.xlsx"
        shutil.copy2(str(primary), str(output_path))
        print(f"Excelファイルをコピーしました: {output_path}")


if __name__ == "__main__":
    main()
