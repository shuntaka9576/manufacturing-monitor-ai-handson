"""シードスクリプトの検証テスト"""

import sqlite3

import openpyxl
import pytest

from db.seed import DB_PATH, EXCEL_PATH, main


@pytest.fixture(scope="session")
def db_conn():
    """シードスクリプトを実行して DB を生成し、接続を提供する。"""
    main()
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    yield conn
    conn.close()


@pytest.fixture(scope="session")
def excel_wb():
    """Excel ワークブックを読み込んで提供する。"""
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    yield wb
    wb.close()


# --- レコード件数と基本制約のテスト ---


def test_record_counts(db_conn):
    """各テーブルのレコード件数が期待値と一致することを確認する。"""
    cur = db_conn.cursor()
    assert cur.execute("SELECT COUNT(*) FROM equipment").fetchone()[0] == 8
    assert cur.execute("SELECT COUNT(*) FROM status_logs").fetchone()[0] == 59
    assert cur.execute("SELECT COUNT(*) FROM sensor_readings").fetchone()[0] == 1152


def test_equipment_ids_sequential(db_conn):
    """設備IDが1〜8の連番であることを確認する。"""
    cur = db_conn.cursor()
    ids = [row[0] for row in cur.execute("SELECT id FROM equipment ORDER BY id").fetchall()]
    assert ids == list(range(1, 9))


def test_foreign_key_constraints(db_conn):
    """外部キー制約が正しく機能していることを確認する。"""
    cur = db_conn.cursor()
    with pytest.raises(sqlite3.IntegrityError):
        cur.execute(
            "INSERT INTO status_logs "
            "(equipment_id, timestamp, old_status, new_status, reason) "
            "VALUES (999, '2026-01-01T00:00:00', '稼働中', '停止中', 'テスト')"
        )
    db_conn.rollback()


def test_cnc_rpm_not_null(db_conn):
    """CNC旋盤（設備ID 1, 2）の rpm が NULL でないことを確認する。"""
    cur = db_conn.cursor()
    null_count = cur.execute(
        "SELECT COUNT(*) FROM sensor_readings WHERE equipment_id IN (1, 2) AND rpm IS NULL"
    ).fetchone()[0]
    assert null_count == 0


# --- ラウンドトリップ整合性テスト ---


def test_equipment_roundtrip(db_conn, excel_wb):
    """Excel の設備マスタ全行と equipment テーブル全行が一致することを確認する。"""
    ws = excel_wb["設備マスタ"]
    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, equipment_type, location, installed_date = row
        if hasattr(installed_date, "strftime"):
            installed_date = installed_date.strftime("%Y-%m-%d")
        else:
            installed_date = str(installed_date)
        excel_rows.append((name, equipment_type, location, installed_date))

    cur = db_conn.cursor()
    db_rows = cur.execute("SELECT name, equipment_type, location, installed_date FROM equipment ORDER BY id").fetchall()

    assert len(excel_rows) == len(db_rows)
    for excel_row, db_row in zip(excel_rows, db_rows, strict=True):
        assert excel_row == db_row


def test_status_logs_roundtrip(db_conn, excel_wb):
    """Excel のステータス変更履歴全行と status_logs テーブル全行が一致することを確認する。"""
    ws = excel_wb["ステータス変更履歴"]
    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            equipment_id,
            _equipment_name,
            timestamp,
            old_status,
            new_status,
            reason,
        ) = row
        if hasattr(timestamp, "isoformat"):
            timestamp = timestamp.isoformat()
        else:
            timestamp = str(timestamp)
        excel_rows.append(
            (
                int(equipment_id),
                timestamp,
                old_status,
                new_status,
                reason,
            )
        )

    cur = db_conn.cursor()
    db_rows = cur.execute(
        "SELECT equipment_id, timestamp, old_status, new_status, reason FROM status_logs ORDER BY id"
    ).fetchall()

    assert len(excel_rows) == len(db_rows)
    for excel_row, db_row in zip(excel_rows, db_rows, strict=True):
        assert excel_row == db_row


def test_sensor_readings_roundtrip(db_conn, excel_wb):
    """Excel のセンサーデータ全行と sensor_readings テーブル全行が一致することを確認する（NaN は NULL として比較）。"""
    import math

    ws = excel_wb["センサーデータ"]
    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure = row
        if hasattr(timestamp, "isoformat"):
            timestamp = timestamp.isoformat()
        else:
            timestamp = str(timestamp)

        def to_none(v):
            if v is None:
                return None
            if isinstance(v, float) and math.isnan(v):
                return None
            return v

        excel_rows.append(
            (
                int(equipment_id),
                timestamp,
                to_none(temperature),
                to_none(vibration),
                to_none(rpm),
                to_none(power_kw),
                to_none(pressure),
            )
        )

    cur = db_conn.cursor()
    db_rows = cur.execute(
        "SELECT equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure "
        "FROM sensor_readings ORDER BY id"
    ).fetchall()

    assert len(excel_rows) == len(db_rows)
    for excel_row, db_row in zip(excel_rows, db_rows, strict=True):
        assert excel_row == db_row


def test_equipment_status_updated(db_conn):
    """update_equipment_status() により equipment.status が正しく更新されていることを確認する。"""
    cur = db_conn.cursor()
    rows = cur.execute(
        "SELECT e.id, e.status, "
        "(SELECT sl.new_status FROM status_logs sl "
        " WHERE sl.equipment_id = e.id ORDER BY sl.timestamp DESC LIMIT 1) "
        "FROM equipment e ORDER BY e.id"
    ).fetchall()

    for equip_id, actual_status, expected_status in rows:
        if expected_status is not None:
            assert actual_status == expected_status, (
                f"設備ID {equip_id}: status={actual_status}, expected={expected_status}"
            )
