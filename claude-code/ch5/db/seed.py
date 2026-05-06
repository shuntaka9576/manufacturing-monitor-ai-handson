import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import logging
import sqlite3
from datetime import datetime

from openpyxl import load_workbook

from db.connection import DB_PATH, apply_schema, connect

XLSX_PATH = Path(__file__).resolve().parent.parent / "sample_data.xlsx"

REQUIRED_SHEETS = {"設備マスタ", "センサーデータ", "ステータス変更履歴"}

EQUIPMENT_COLS = {"設備名", "タイプ", "設置場所", "設置日"}
SENSOR_COLS = {"設備ID", "タイムスタンプ", "temperature", "vibration", "rpm", "power_kw", "pressure"}
STATUS_COLS = {"設備ID", "発生日時", "変更前ステータス", "変更後ステータス", "理由"}

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)


def _extract_equipment(ws) -> list[tuple]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    idx = {c: headers.index(c) for c in EQUIPMENT_COLS}
    records = []
    for eq_id, row in enumerate(rows[1:], start=1):
        name = row[idx["設備名"]]
        type_ = row[idx["タイプ"]]
        location = row[idx["設置場所"]]
        installed_on_raw = row[idx["設置日"]]
        for field, val in [("設備名", name), ("タイプ", type_), ("設置場所", location), ("設置日", installed_on_raw)]:
            if val is None:
                raise ValueError(f"設備マスタ row {eq_id}: 必須セル '{field}' が空")
        if isinstance(installed_on_raw, datetime):
            installed_on = installed_on_raw.date().isoformat()
        else:
            datetime.fromisoformat(str(installed_on_raw))
            installed_on = str(installed_on_raw)
        records.append((eq_id, name, type_, location, installed_on))
    return records


def _extract_sensor_readings(ws, valid_eq_ids: set) -> list[tuple]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    idx = {c: headers.index(c) for c in SENSOR_COLS}
    records = []
    for row_num, row in enumerate(rows[1:], start=2):
        eq_id = row[idx["設備ID"]]
        ts_raw = row[idx["タイムスタンプ"]]
        temperature = row[idx["temperature"]]
        vibration = row[idx["vibration"]]
        rpm = row[idx["rpm"]]
        power_kw = row[idx["power_kw"]]
        pressure = row[idx["pressure"]]

        for field, val in [
            ("設備ID", eq_id),
            ("タイムスタンプ", ts_raw),
            ("temperature", temperature),
            ("vibration", vibration),
            ("power_kw", power_kw),
        ]:
            if val is None:
                raise ValueError(f"センサーデータ row {row_num}: 必須セル '{field}' が空")

        if eq_id not in valid_eq_ids:
            raise ValueError(f"センサーデータ row {row_num}: 設備ID={eq_id} が設備マスタに存在しない (FK 違反)")

        if isinstance(ts_raw, datetime):
            ts = ts_raw.isoformat()
        else:
            datetime.fromisoformat(str(ts_raw))
            ts = str(ts_raw)

        temperature = float(temperature)
        vibration = float(vibration)
        rpm = float(rpm) if rpm is not None else None
        power_kw = float(power_kw)
        pressure = float(pressure) if pressure is not None else None

        records.append((int(eq_id), ts, temperature, vibration, rpm, power_kw, pressure))
    return records


def _extract_status_logs(ws, valid_eq_ids: set) -> list[tuple]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    idx = {c: headers.index(c) for c in STATUS_COLS}
    records = []
    for row_num, row in enumerate(rows[1:], start=2):
        eq_id = row[idx["設備ID"]]
        occurred_at_raw = row[idx["発生日時"]]
        prev_status = row[idx["変更前ステータス"]]
        new_status = row[idx["変更後ステータス"]]
        reason = row[idx["理由"]]

        for field, val in [
            ("設備ID", eq_id),
            ("発生日時", occurred_at_raw),
            ("変更前ステータス", prev_status),
            ("変更後ステータス", new_status),
            ("理由", reason),
        ]:
            if val is None:
                raise ValueError(f"ステータス変更履歴 row {row_num}: 必須セル '{field}' が空")

        if eq_id not in valid_eq_ids:
            raise ValueError(f"ステータス変更履歴 row {row_num}: 設備ID={eq_id} が設備マスタに存在しない (FK 違反)")

        if isinstance(occurred_at_raw, datetime):
            occurred_at = occurred_at_raw.isoformat()
        else:
            datetime.fromisoformat(str(occurred_at_raw))
            occurred_at = str(occurred_at_raw)

        records.append((int(eq_id), occurred_at, str(prev_status), str(new_status), str(reason)))
    return records


def main() -> int:
    try:
        if not XLSX_PATH.exists():
            logger.error("sample_data.xlsx not found: %s", XLSX_PATH)
            return 2

        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
        missing_sheets = REQUIRED_SHEETS - set(wb.sheetnames)
        if missing_sheets:
            logger.error("必須シートが見つかりません: %s", missing_sheets)
            return 1

        for sheet_name, required_cols in [
            ("設備マスタ", EQUIPMENT_COLS),
            ("センサーデータ", SENSOR_COLS),
            ("ステータス変更履歴", STATUS_COLS),
        ]:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = set(str(c) for c in header_row if c is not None)
            missing_cols = required_cols - headers
            if missing_cols:
                logger.error("シート '%s' に必須列がありません: %s", sheet_name, missing_cols)
                return 1

        wb.close()

        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
        equipment = _extract_equipment(wb["設備マスタ"])
        valid_eq_ids = {row[0] for row in equipment}
        sensor_readings = _extract_sensor_readings(wb["センサーデータ"], valid_eq_ids)
        status_logs = _extract_status_logs(wb["ステータス変更履歴"], valid_eq_ids)
        wb.close()

        conn = connect()
        apply_schema(conn)
        logger.info("schema applied: %s", DB_PATH)

        with conn:
            conn.execute("DELETE FROM sensor_readings")
            conn.execute("DELETE FROM status_logs")
            conn.execute("DELETE FROM equipment")

            conn.executemany(
                "INSERT INTO equipment (equipment_id, name, type, location, installed_on) VALUES (?, ?, ?, ?, ?)",
                equipment,
            )

            sr_seen: set[tuple] = set()
            sr_dups = 0
            for rec in sensor_readings:
                key = (rec[0], rec[1])
                if key in sr_seen:
                    logger.warning("duplicate sensor_readings key: equipment_id=%s, ts=%s", key[0], key[1])
                    sr_dups += 1
                sr_seen.add(key)
                conn.execute(
                    "INSERT OR REPLACE INTO sensor_readings VALUES (?, ?, ?, ?, ?, ?, ?)",
                    rec,
                )

            sl_seen: set[tuple] = set()
            sl_dups = 0
            for rec in status_logs:
                key = (rec[0], rec[1])
                if key in sl_seen:
                    logger.warning("duplicate status_logs key: equipment_id=%s, occurred_at=%s", key[0], key[1])
                    sl_dups += 1
                sl_seen.add(key)
                conn.execute(
                    "INSERT OR REPLACE INTO status_logs VALUES (?, ?, ?, ?, ?)",
                    rec,
                )

            if sr_dups or sl_dups:
                logger.info(
                    "duplicates detected (last-write-wins): sensor_readings=%d, status_logs=%d",
                    sr_dups,
                    sl_dups,
                )

            conn.execute(
                """
                UPDATE equipment
                SET status = (
                    SELECT new_status FROM status_logs sl
                    WHERE sl.equipment_id = equipment.equipment_id
                    ORDER BY sl.occurred_at DESC LIMIT 1
                )
                WHERE EXISTS (
                    SELECT 1 FROM status_logs sl
                    WHERE sl.equipment_id = equipment.equipment_id
                )
                """
            )

        logger.info("loaded sheets from %s", XLSX_PATH)

        n_eq = conn.execute("SELECT COUNT(*) FROM equipment").fetchone()[0]
        n_sr = conn.execute("SELECT COUNT(*) FROM sensor_readings").fetchone()[0]
        n_sl = conn.execute("SELECT COUNT(*) FROM status_logs").fetchone()[0]
        conn.close()

        print(f"equipment={n_eq}, sensor_readings={n_sr}, status_logs={n_sl}")
        return 0

    except FileNotFoundError as e:
        logger.error("seed aborted: %s", e)
        return 2
    except (ValueError, sqlite3.IntegrityError) as e:
        logger.error("seed aborted: %s", e)
        return 1


if __name__ == "__main__":
    import sys

    sys.exit(main())
