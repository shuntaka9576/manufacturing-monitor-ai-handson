"""シードスクリプトの検証テスト"""

import math
import sqlite3

import openpyxl
import pytest

from db.connection import DB_PATH
from db.seed import XLSX_PATH, main


@pytest.fixture(scope="session")
def db_conn():
    """シードスクリプトを実行して DB を生成し、接続を提供する。"""
    rc = main()
    assert rc == 0, f"seed main() returned non-zero: {rc}"
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    yield conn
    conn.close()


@pytest.fixture(scope="session")
def excel_wb():
    """Excel ワークブックを読み込んで提供する。"""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    yield wb
    wb.close()


def _to_iso(ts):
    if hasattr(ts, "isoformat"):
        return ts.isoformat()
    return str(ts)


def _to_date(d):
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _nan_to_none(v):
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    return v


# --- レコード件数と基本制約のテスト ---


def test_record_counts(db_conn):
    """各テーブルのレコード件数が期待値と一致することを確認する。"""
    cur = db_conn.cursor()
    assert cur.execute("SELECT COUNT(*) FROM equipment").fetchone()[0] == 8
    assert cur.execute("SELECT COUNT(*) FROM status_logs").fetchone()[0] == 59
    assert cur.execute("SELECT COUNT(*) FROM sensor_readings").fetchone()[0] == 1152


def test_equipment_ids_sequential(db_conn):
    """設備IDが1〜8の連番であることを確認する。"""
    cur = db_conn.cursor()
    ids = [
        row[0]
        for row in cur.execute(
            "SELECT equipment_id FROM equipment ORDER BY equipment_id"
        ).fetchall()
    ]
    assert ids == list(range(1, 9))


def test_foreign_key_constraints(db_conn):
    """外部キー制約が正しく機能していることを確認する。"""
    cur = db_conn.cursor()
    with pytest.raises(sqlite3.IntegrityError):
        cur.execute(
            "INSERT INTO status_logs "
            "(equipment_id, occurred_at, prev_status, new_status, reason) "
            "VALUES (999, '2026-01-01T00:00:00', '稼働中', '停止中', 'テスト')"
        )
    db_conn.rollback()


def test_cnc_rpm_not_null(db_conn):
    """CNC旋盤（設備ID 1, 2）の rpm が NULL でないことを確認する。"""
    cur = db_conn.cursor()
    null_count = cur.execute(
        "SELECT COUNT(*) FROM sensor_readings WHERE equipment_id IN (1, 2) AND rpm IS NULL"
    ).fetchone()[0]
    assert null_count == 0


# --- ラウンドトリップ整合性テスト ---


def test_equipment_roundtrip(db_conn, excel_wb):
    """Excel の設備マスタ全行と equipment テーブル全行が一致することを確認する。"""
    ws = excel_wb["設備マスタ"]
    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, equipment_type, location, installed_on = row[:4]
        excel_rows.append((name, equipment_type, location, _to_date(installed_on)))

    cur = db_conn.cursor()
    db_rows = cur.execute(
        "SELECT name, type, location, installed_on FROM equipment ORDER BY equipment_id"
    ).fetchall()

    assert len(excel_rows) == len(db_rows)
    for excel_row, db_row in zip(excel_rows, db_rows):
        assert excel_row == db_row


def test_status_logs_roundtrip(db_conn, excel_wb):
    """Excel のステータス変更履歴全行と status_logs テーブル全行が一致することを確認する。"""
    ws = excel_wb["ステータス変更履歴"]
    excel_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        equipment_id, _equipment_name, occurred_at, prev_status, new_status, reason = row
        key = (int(equipment_id), _to_iso(occurred_at))
        excel_map[key] = (prev_status, new_status, reason)

    cur = db_conn.cursor()
    db_map = {
        (eq_id, occurred_at): (prev_s, new_s, reason)
        for eq_id, occurred_at, prev_s, new_s, reason in cur.execute(
            "SELECT equipment_id, occurred_at, prev_status, new_status, reason FROM status_logs"
        )
    }

    assert excel_map == db_map


def test_sensor_readings_roundtrip(db_conn, excel_wb):
    """Excel のセンサーデータ全行と sensor_readings テーブル全行が一致することを確認する（NaN は NULL として比較）。"""
    ws = excel_wb["センサーデータ"]
    excel_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure = row
        key = (int(equipment_id), _to_iso(timestamp))
        excel_map[key] = (
            _nan_to_none(temperature),
            _nan_to_none(vibration),
            _nan_to_none(rpm),
            _nan_to_none(power_kw),
            _nan_to_none(pressure),
        )

    cur = db_conn.cursor()
    db_map = {
        (eq_id, ts): (temp, vib, rpm, power, pressure)
        for eq_id, ts, temp, vib, rpm, power, pressure in cur.execute(
            "SELECT equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure "
            "FROM sensor_readings"
        )
    }

    assert excel_map == db_map


def test_equipment_status_updated(db_conn):
    """equipment.status が status_logs の最新エントリ（occurred_at 降順 1 件）と一致することを確認する。"""
    cur = db_conn.cursor()
    rows = cur.execute(
        "SELECT e.equipment_id, e.status, "
        "(SELECT sl.new_status FROM status_logs sl "
        " WHERE sl.equipment_id = e.equipment_id ORDER BY sl.occurred_at DESC LIMIT 1) "
        "FROM equipment e ORDER BY e.equipment_id"
    ).fetchall()

    for equip_id, actual_status, expected_status in rows:
        if expected_status is not None:
            assert actual_status == expected_status, (
                f"設備ID {equip_id}: status={actual_status}, expected={expected_status}"
            )
