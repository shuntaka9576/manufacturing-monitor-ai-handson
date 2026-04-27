"""製造設備監視ダッシュボード シードスクリプト

Excel ファイル（sample_data.xlsx）から SQLite データベースへ
シードデータを投入する。
"""

import math
import sqlite3
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# パス定数
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "sample_data.xlsx"
DB_PATH = PROJECT_ROOT / "db" / "equipment_monitoring.db"
SCHEMA_PATH = PROJECT_ROOT / "db" / "schema.sql"

# 期待するシート名
EXPECTED_SHEETS = {"設備マスタ", "ステータス変更履歴", "センサーデータ"}


def to_sql_value(value):
    """Excel セル値を SQL 投入用の値に変換する。

    NaN を None に変換し、SQL の NULL として扱えるようにする。
    """
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def validate_excel(excel_path: Path) -> Workbook:
    """Excel ファイルの存在確認とシート名検証を行う。

    Args:
        excel_path: Excel ファイルのパス

    Returns:
        openpyxl.Workbook オブジェクト

    Raises:
        SystemExit: ファイルが存在しない、またはシート名が期待と異なる場合
    """
    if not excel_path.exists():
        raise SystemExit(f"エラー: Excel ファイルが見つかりません: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    actual_sheets = set(wb.sheetnames)

    missing = EXPECTED_SHEETS - actual_sheets
    if missing:
        wb.close()
        raise SystemExit(
            f"エラー: 必要なシートが見つかりません: {', '.join(sorted(missing))}"
        )

    return wb


def create_database(db_path: Path, schema_path: Path) -> sqlite3.Connection:
    """既存DBを削除し、新規作成してスキーマを適用する。

    Args:
        db_path: SQLite データベースファイルのパス
        schema_path: schema.sql ファイルのパス

    Returns:
        sqlite3.Connection オブジェクト
    """
    # 既存DBファイルが存在する場合は削除
    if db_path.exists():
        db_path.unlink()

    # 新規DB作成 & スキーマ適用
    conn = sqlite3.connect(str(db_path))
    schema_sql = schema_path.read_text(encoding="utf-8")
    conn.executescript(schema_sql)

    # 外部キー制約を有効化
    conn.execute("PRAGMA foreign_keys = ON")

    return conn


def insert_equipment(ws, cursor: sqlite3.Cursor) -> int:
    """設備マスタシートの全行を読み込み、equipment テーブルへ投入する。

    行インデックス（0始まり）+ 1 を id として設定する。

    Args:
        ws: 設備マスタの Worksheet オブジェクト
        cursor: sqlite3.Cursor オブジェクト

    Returns:
        投入件数
    """
    count = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        name, type_, location, installation_date = row
        # datetime オブジェクトの場合は YYYY-MM-DD 文字列に変換
        if hasattr(installation_date, "strftime"):
            installation_date = installation_date.strftime("%Y-%m-%d")
        else:
            installation_date = str(installation_date)
        equipment_id = idx + 1
        cursor.execute(
            "INSERT INTO equipment (id, name, type, location, installation_date) "
            "VALUES (?, ?, ?, ?, ?)",
            (equipment_id, name, type_, location, installation_date),
        )
        count += 1
    return count


def insert_status_history(ws, cursor: sqlite3.Cursor) -> int:
    """ステータス変更履歴シートの全行を読み込み、status_history テーブルへ投入する。

    Args:
        ws: ステータス変更履歴の Worksheet オブジェクト
        cursor: sqlite3.Cursor オブジェクト

    Returns:
        投入件数
    """
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            equipment_id,
            equipment_name,
            occurred_at,
            status_before,
            status_after,
            reason,
        ) = row
        # datetime オブジェクトの場合は ISO8601 文字列に変換
        if hasattr(occurred_at, "isoformat"):
            occurred_at = occurred_at.isoformat()
        else:
            occurred_at = str(occurred_at)
        cursor.execute(
            "INSERT INTO status_history "
            "(equipment_id, equipment_name, occurred_at, status_before, status_after, reason) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                int(equipment_id),
                equipment_name,
                occurred_at,
                status_before,
                status_after,
                reason,
            ),
        )
        count += 1
    return count


def insert_sensor_data(ws, cursor: sqlite3.Cursor) -> int:
    """センサーデータシートの全行を読み込み、sensor_data テーブルへ投入する。

    NaN 値は to_sql_value() で None に変換し、SQL の NULL として投入する。

    Args:
        ws: センサーデータの Worksheet オブジェクト
        cursor: sqlite3.Cursor オブジェクト

    Returns:
        投入件数
    """
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure = row
        # datetime オブジェクトの場合は ISO8601 文字列に変換
        if hasattr(timestamp, "isoformat"):
            timestamp = timestamp.isoformat()
        else:
            timestamp = str(timestamp)
        cursor.execute(
            "INSERT INTO sensor_data "
            "(equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                int(equipment_id),
                timestamp,
                to_sql_value(temperature),
                to_sql_value(vibration),
                to_sql_value(rpm),
                to_sql_value(power_kw),
                to_sql_value(pressure),
            ),
        )
        count += 1
    return count


def load_and_insert_data(wb: Workbook, conn: sqlite3.Connection) -> dict[str, int]:
    """単一トランザクション内で3テーブルへデータを投入する。

    Args:
        wb: openpyxl.Workbook オブジェクト
        conn: sqlite3.Connection オブジェクト

    Returns:
        各テーブルの投入件数: {"equipment": N, "status_history": N, "sensor_data": N}

    Raises:
        エラー発生時はロールバックし、例外を再送出する
    """
    try:
        cursor = conn.cursor()
        eq_count = insert_equipment(wb["設備マスタ"], cursor)
        sh_count = insert_status_history(wb["ステータス変更履歴"], cursor)
        sd_count = insert_sensor_data(wb["センサーデータ"], cursor)
        conn.commit()
    except Exception:
        conn.rollback()
        raise

    return {"equipment": eq_count, "status_history": sh_count, "sensor_data": sd_count}


def print_summary(counts: dict[str, int]) -> None:
    """各テーブルのレコード件数と完了メッセージを標準出力に表示する。

    Args:
        counts: テーブル名をキー、件数を値とする辞書
    """
    print(f"設備マスタ: {counts['equipment']}件")
    print(f"ステータス変更履歴: {counts['status_history']}件")
    print(f"センサーデータ: {counts['sensor_data']}件")
    print("シード完了")


def main() -> None:
    """エントリーポイント。Excel からデータを読み込み SQLite へ投入する。"""
    wb = validate_excel(EXCEL_PATH)
    conn = create_database(DB_PATH, SCHEMA_PATH)
    try:
        counts = load_and_insert_data(wb, conn)
        print_summary(counts)
    finally:
        wb.close()
        conn.close()


if __name__ == "__main__":
    main()
