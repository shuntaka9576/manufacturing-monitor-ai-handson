"""製造設備モニタリング用シードデータ投入スクリプト

sample_data.xlsx を読み込み、SQLiteデータベースにデータを投入する。
全てのデータを Excel から取得し、変換や生成は行わない。
"""

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

DB_PATH = Path(__file__).parent.parent / "data" / "factory.db"
SCHEMA_PATH = Path(__file__).parent / "schema.sql"
EXCEL_PATH = Path(__file__).parent.parent / "sample_data.xlsx"


def load_equipment_data(wb):
    """設備マスタシートから設備データを読み込む"""
    ws = wb["設備マスタ"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        rows.append(tuple(row))
    return rows


def load_status_log_events(wb):
    """ステータス変更履歴シートからイベントを読み込む

    Returns:
        [(equip_id, timestamp, old_status, new_status, reason), ...] のリスト
    """
    ws = wb["ステータス変更履歴"]
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        equip_id, _name, timestamp, old_s, new_s, reason = row
        ts = timestamp if isinstance(timestamp, str) else timestamp.isoformat()
        events.append((int(equip_id), ts, old_s, new_s, reason))
    return events


def load_sensor_data(wb):
    """センサーデータシートから時系列データを読み込む

    Returns:
        [(equip_id, timestamp, temperature, vibration, rpm, power_kw, pressure), ...] のリスト
    """
    ws = wb["センサーデータ"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        equip_id, timestamp, temp, vib, rpm, power, pressure = row
        ts = timestamp if isinstance(timestamp, str) else timestamp.isoformat()
        rows.append((int(equip_id), ts, temp, vib, rpm, power, pressure))
    return rows


def create_tables(conn: sqlite3.Connection) -> None:
    schema = SCHEMA_PATH.read_text(encoding="utf-8")
    conn.executescript(schema)


def seed_equipment(conn: sqlite3.Connection, equipment_data: list) -> None:
    conn.executemany(
        "INSERT INTO equipment (name, equipment_type, location, installed_date) VALUES (?, ?, ?, ?)",
        equipment_data,
    )
    conn.commit()


def seed_status_logs(conn: sqlite3.Connection, events: list) -> None:
    conn.executemany(
        "INSERT INTO status_logs (equipment_id, timestamp, old_status, new_status, reason) VALUES (?, ?, ?, ?, ?)",
        events,
    )
    conn.commit()


def seed_sensor_readings(conn: sqlite3.Connection, sensor_data: list) -> None:
    conn.executemany(
        "INSERT INTO sensor_readings"
        " (equipment_id, timestamp, temperature, vibration, rpm, power_kw, pressure)"
        " VALUES (?, ?, ?, ?, ?, ?, ?)",
        sensor_data,
    )
    conn.commit()


def update_equipment_status(conn: sqlite3.Connection) -> None:
    """ステータス変更履歴の最新エントリから各設備のステータスを導出して更新する"""
    conn.execute("""
        UPDATE equipment
        SET status = (
            SELECT sl.new_status
            FROM status_logs sl
            WHERE sl.equipment_id = equipment.id
            ORDER BY sl.timestamp DESC
            LIMIT 1
        )
        WHERE EXISTS (
            SELECT 1 FROM status_logs sl WHERE sl.equipment_id = equipment.id
        )
    """)
    conn.commit()


def main() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        DB_PATH.unlink()

    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    equipment_data = load_equipment_data(wb)
    status_log_events = load_status_log_events(wb)
    sensor_data = load_sensor_data(wb)
    wb.close()

    conn = sqlite3.connect(str(DB_PATH))
    create_tables(conn)
    seed_equipment(conn, equipment_data)
    seed_status_logs(conn, status_log_events)
    seed_sensor_readings(conn, sensor_data)
    update_equipment_status(conn)
    conn.close()
    print(f"データベースを作成しました: {DB_PATH}")


if __name__ == "__main__":
    main()
